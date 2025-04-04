import tkinter as tk
from tkinter import filedialog, Label, Button, messagebox
import time
import threading
import queue
import os
import pandas as pd  # Added missing import
import xlwings as xw  
import itertools
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter




def processing_function(df_wire_list, mmsta_sep, PN_columns):
      all_idx_integrated = []

      for wire in df_wire_list['Wire Internal Name'].unique():
            df_wire = df_wire_list[df_wire_list['Wire Internal Name'] == wire]
            matching_rows = mmsta_sep[mmsta_sep['Wire Internal Name'] == wire]
            if df_wire.shape[0] == 1 and not matching_rows.empty:
                  pene_max = df_wire[PN_columns].values[0]
                  pen_mmsta = np.sum(matching_rows[PN_columns].values, axis=0)
                  if np.array_equal(pene_max, pen_mmsta):
                        n_fois_duplicate = mmsta_sep[mmsta_sep['Wire Internal Name'] == wire].shape[0]

                        if n_fois_duplicate:
                  
                              row_to_duplicate = df_wire.copy()
                              df_wire_list = df_wire_list[df_wire_list['Wire Internal Name'] != wire]
                              # df_wire_list = df_wire_list.reset_index(drop=True)
                        
                              for i in range(n_fois_duplicate):
                                    matching_row = matching_rows.iloc[i]
                                    level = matching_row['Level'][-1]
                                    row_to_duplicate[PN_columns] = matching_row[PN_columns].values
                                    if int(level) == 3:
                                          row_to_duplicate['SN1'] = matching_row['SN3'] 
                                          row_to_duplicate['SN2'] = matching_row['SN2']
                                          row_to_duplicate['Sup Grp'] = matching_row['SN1']
                                          row_to_duplicate['DES'] = matching_row['DES3']
                                          row_to_duplicate['TYPE'] = matching_row['TYPE']
                                    elif int(level) == 2:
                                          row_to_duplicate['SN1'] = matching_row['SN2']
                                          row_to_duplicate['SN2'] = matching_row['SN1']
                                          row_to_duplicate['DES'] = matching_row['DES2']
                                          row_to_duplicate['TYPE'] = matching_row['TYPE']
                                    elif int(level) == 1:
                                          row_to_duplicate['SN1'] = matching_row['SN1']
                                          row_to_duplicate['DES'] = matching_row['DES1']
                                          row_to_duplicate['TYPE'] = matching_row['TYPE']
                                          
                                    all_idx_integrated.extend(matching_rows.index)
                                    df_wire_list = pd.concat([df_wire_list, row_to_duplicate])
                  

                              
            elif df_wire.shape[0] > 1 and not matching_rows.empty:
                  if matching_rows.shape[0] >= df_wire.shape[0]: #matching_rows.shape[0] > df_wire.shape[0]:
                        indices_valides = get_idx(df_wire,matching_rows, PN_columns) #(idx_max_wire,(les indices de mmsta))
                        if indices_valides :
                              for idx_val in indices_valides :
                                    n_fois_duplicate = len(idx_val[1])
                                    idx_mmsta = idx_val[1]
                                    idx_row_to_duplicate = idx_val[0]
                                    if n_fois_duplicate >= 1 :
                                          row_to_duplicate = df_wire.loc[idx_row_to_duplicate]
                                          df_wire_list = df_wire_list.drop(idx_row_to_duplicate)
                                          # df_wire_list = df_wire_list.reset_index(drop=True)
                                    
                                          for i in list(idx_mmsta):
                                                matching_row = matching_rows.loc[i]
                                                level = matching_row['Level'][-1]
                                                row_to_duplicate[PN_columns] = matching_row[PN_columns].values
                                                if int(level) == 3:
                                                      row_to_duplicate['SN1'] = matching_row['SN3'] 
                                                      row_to_duplicate['SN2'] = matching_row['SN2']
                                                      row_to_duplicate['Sup Grp'] = matching_row['SN1']
                                                      row_to_duplicate['DES'] = matching_row['DES3']
                                                      row_to_duplicate['TYPE'] = matching_row['TYPE']
                                                elif int(level) == 2:
                                                      row_to_duplicate['SN1'] = matching_row['SN2']
                                                      row_to_duplicate['SN2'] = matching_row['SN1']
                                                      row_to_duplicate['DES'] = matching_row['DES2']
                                                      row_to_duplicate['TYPE'] = matching_row['TYPE']
                                                elif int(level) == 1:
                                                      row_to_duplicate['SN1'] = matching_row['SN1']
                                                      row_to_duplicate['DES'] = matching_row['DES1']
                                                      row_to_duplicate['TYPE'] = matching_row['TYPE']

                                                all_idx_integrated.extend(matching_rows.index)
                                                df_wire_list = pd.concat([df_wire_list, row_to_duplicate.to_frame().T])
                                                # df_wire_list = df_wire_list.reset_index(drop=True)
                        
                        
                  # else:
                  #       tuple_match = find_matching_rows_for_wire(wire, df_wire_list, mmsta_sep, PN_columns_wirelist, PN_columns_mmsta)
                  #       for wire_idx, mmsta_idx in tuple_match:
                  #             matching_row = mmsta_sep.loc[mmsta_idx]
                  #             # 127 __ 956 __ 151
                  #             # 128 __ 574 __ 151
                  #             # # Create a new row based on the matching row
                  #             new_row = df_wire_list.loc[wire_idx]
                  #             df_wire_list = df_wire_list.drop(wire_idx)#df_wire_list[df_wire_list['Wire Internal Name'] != wire]
                  #             # df_wire_list = df_wire_list.reset_index(drop=True)
                  #             new_row[PN_columns_wirelist] = matching_row[PN_columns_mmsta].values
                              
                  #             level = matching_row['Level'][-1]
                              
                  #             if int(level) == 3:
                  #                   new_row['SN1'] = matching_row['SN3'] 
                  #                   new_row['SN2'] = matching_row['SN2']
                  #                   new_row['Sup Grp'] = matching_row['SN1']
                  #                   new_row['DES'] = matching_row['DES3']
                  #                   new_row['SN1'] = matching_row['SN2']
                  #                   new_row['SN2'] = matching_row['SN1']
                  #                   new_row['DES'] = matching_row['DES2']
                  #             elif int(level) == 1:
                  #                   new_row['SN1'] = matching_row['SN1']
                  #                   new_row['DES'] = matching_row['DES1']
                              
                  #             # Append the new row to df_wire_list
                  #             df_wire_list = pd.concat([df_wire_list, new_row.to_frame().T])
      return df_wire_list, all_idx_integrated



def get_col(wire, material_to_comp_dict, mmsta_sep):
     cols = []
     keys_for_Wire = [key for key, value in material_to_comp_dict.items() if value == wire]
     for key in keys_for_Wire :
          col = mmsta_sep.columns[mmsta_sep.isin([key]).any()][0]
          cols.append(col)

     return cols

# Define a function to find rows in df_wire_list that match product number patterns in mmsta_sep
def find_matching_rows_for_wire(wire_internal_name, df_wire_list, mmsta_sep, PN_columns):

    matches = []
    
    # Get rows for the specified wire
    wire_rows = df_wire_list[df_wire_list['Wire Internal Name'] == wire_internal_name]
    matching_rows = mmsta_sep[mmsta_sep['Wire Internal Name'] == wire_internal_name]
    
    if not matching_rows.empty and not wire_rows.empty:
        for wire_idx in wire_rows.index:
            pene_max = wire_rows.loc[wire_idx, PN_columns].values
            
            for mmsta_idx in matching_rows.index:
                pen_mmsta = matching_rows.loc[mmsta_idx, PN_columns].values
                
                # Check if product number patterns match
                if np.array_equal(pene_max, pen_mmsta):
                    matches.append((wire_idx, mmsta_idx))
    
    return matches



def get_idx(df_wire,matching_rows, PN_columns):
    indices_valides = []
    for idx in df_wire.index:
        target = df_wire.loc[idx, PN_columns].values
        for r in range(1, len(matching_rows) + 1):  
            for indices in itertools.combinations(matching_rows.index, r):
                pene = matching_rows.loc[list(indices)][PN_columns].values
                sum_pene = np.sum(pene, axis=0)
                if np.array_equal(sum_pene, target):
                    indices_valides.append((idx,indices))

    #print("Indices des lignes valides :", indices_valides)
    return indices_valides

def Add_TYPE_col(mmsta_sep):
    mmsta_sep.loc[mmsta_sep['DES1'].str.contains('Circuit', na=False), 'TYPE'] = 'Circuit'
    mmsta_sep.loc[mmsta_sep['DES1'].str.contains('Double', na=False), 'TYPE'] = 'Double'
    mmsta_sep.loc[mmsta_sep['DES1'].str.contains('Twisted', na=False), 'TYPE'] = 'Twisted'
    mmsta_sep.loc[mmsta_sep['DES1'].str.contains('Joint', na=False), 'TYPE'] = 'Joint'
    mmsta_sep.loc[mmsta_sep['DES1'].str.contains('Super Group', na=False), 'TYPE'] = 'Super Group'

    return mmsta_sep

def get_idx_duplicate(df_wire,matching_rows, PN_columns):
    indices_valides = []
    for idx in df_wire.index:
        target = df_wire.loc[idx, PN_columns].values
        for r in range(1, len(matching_rows) + 1):  
            for indices in itertools.combinations(matching_rows.index, r):
                pene = matching_rows.loc[list(indices)][PN_columns].values
                sum_pene = np.sum(pene, axis=0)
                if np.array_equal(sum_pene, target):
                    indices_valides.append((df_wire.loc[idx,'Composition No'],indices))

    #print("Indices des lignes valides :", indices_valides)
    return indices_valides

def check_duplicate(ypp):
    duplicate_materials = ypp['Material'].value_counts()
    duplicates = duplicate_materials[duplicate_materials > 1]
    sn_duplicated = duplicates.index.tolist()
    return sn_duplicated


def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

def add_columns(df):
    global cols_to_add 
    cols_to_add = ['PN','SN1','DES1', 'SN2','DES2', 'SN3','DES3']
    col_index = df.columns.get_loc('Material Description') + 1
    for i, col in enumerate(cols_to_add):
        df.insert(col_index + i, col, '')
    return df

def Fill(level : str, df_mmsta):
    for idx in range(df_mmsta.shape[0]):
        if df_mmsta.loc[:,'Level'][idx] == level:
            # PN_arr.append(df_mmsta.loc[idx]['Material Description'])
            df_mmsta.loc[idx,'PN'] = df_mmsta.loc[idx]['Material Description']
            idx_last_l = idx
        else :
            df_mmsta.loc[idx,'PN'] = df_mmsta.loc[idx_last_l]['Material Description']
            # PN_arr.append(df_mmsta.loc[idx_last_l]['Material Description'])
    return df_mmsta

def delete_YFG_YDM(df_mmsta):
    return df_mmsta[~df_mmsta['Material Type'].isin(['YFG', 'YDM'])]

def Fill_with_col(level : str, cols,df):
    idx_last_l = None
    for idx in range(df.shape[0]):
        if df.loc[idx,'Level'] == level:
            df.loc[idx,cols[0]] = df.loc[idx,'Material']
            df.loc[idx,cols[1]] = df.loc[idx,'Material Description']
            idx_last_l = idx
           
            
        else :
            if idx_last_l is not None:
                df.loc[idx,cols[0]] = df.loc[idx_last_l,'Material']
                df.loc[idx,cols[1]] = df.loc[idx_last_l,'Material Description']
                
            else :
                df.loc[idx,cols[0]] = '(Blank)'
                df.loc[idx,cols[1]] = '(Blank)'
    return df

class ExcelProcessor:
    def __init__(self, root):
        self.root = root
        self.MMSTA_file_path = ""
        self.Maxwire_file_path = ""
        self.YPP_file_path = ""
        self.output_dir = ""
        self.processing = False
        self.status_queue = queue.Queue()
        
        self.setup_ui()
        self.check_queue()
    
    def setup_ui(self):
        title_label = Label(self.root, text="Data Integrator Processor", font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        
        # MMSTA file selection
        self.MMSTA_button = Button(self.root, text="Select MMSTA File", command=self.select_MMSTA_file)
        self.MMSTA_button.pack(pady=5)
        self.MMSTA_label = Label(self.root, text="MMSTA: Not selected", wraplength=400)
        self.MMSTA_label.pack()
        
        # Maxwire list file selection
        self.Maxwire_button = Button(self.root, text="Select Maxwire List", command=self.select_Maxwire_file)
        self.Maxwire_button.pack(pady=5)
        self.Maxwire_label = Label(self.root, text="Maxwire List: Not selected", wraplength=400)
        self.Maxwire_label.pack()
        
        # YPP file selection
        self.YPP_button = Button(self.root, text="Select YPP File", command=self.select_YPP_file)
        self.YPP_button.pack(pady=5)
        self.YPP_label = Label(self.root, text="YPP: Not selected", wraplength=400)
        self.YPP_label.pack()
        
        # Output directory selection
        self.output_button = Button(self.root, text="Select Output Folder", command=self.select_output_dir)
        self.output_button.pack(pady=5)
        self.output_label = Label(self.root, text="Output folder: Not selected", wraplength=400)
        self.output_label.pack()
        
        # Process button
        self.process_button = Button(self.root, text="Process Files", command=self.main_separator, 
                              bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), 
                              width=15, height=1, state=tk.DISABLED)
        self.process_button.pack(pady=20)
        
        self.progress_label = Label(self.root, text="", fg="blue")
        self.progress_label.pack()
        
        self.status_label = Label(self.root, text="", fg="black")
        self.status_label.pack()
    
    
    def select_MMSTA_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls;*.XLSX")])
        
        if not file_path:
            return
            
        if not file_path.lower().endswith((".xlsx", ".xls")):
            messagebox.showerror("Error", "Please select a valid Excel file (*.xlsx, *.xls;*.XLSX)")
            return

        self.MMSTA_file_path = file_path
        self.MMSTA_label.config(text=f"MMSTA File: {os.path.basename(self.MMSTA_file_path)}")
        if self.output_dir and self.Maxwire_file_path and self.YPP_file_path: 
            self.process_button.config(state=tk.NORMAL)
    
    def select_Maxwire_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls;*.XLSX")])
        
        if not file_path:
            return
            
        if not file_path.lower().endswith((".xlsx", ".xls")):
            messagebox.showerror("Error", "Please select a valid Excel file (*.xlsx, *.xls;*.XLSX)")
            return

        self.Maxwire_file_path = file_path
        self.Maxwire_label.config(text=f"Maxwire List: {os.path.basename(self.Maxwire_file_path)}")
        if self.output_dir and self.MMSTA_file_path and self.YPP_file_path: 
            self.process_button.config(state=tk.NORMAL)
    
    def select_YPP_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls;*.XLSX")])
        
        if not file_path:
            return
            
        if not file_path.lower().endswith((".xlsx", ".xls")):
            messagebox.showerror("Error", "Please select a valid Excel file (*.xlsx, *.xls;*.XLSX)")
            return

        self.YPP_file_path = file_path
        self.YPP_label.config(text=f"YPP File: {os.path.basename(self.YPP_file_path)}")
        if self.output_dir and self.Maxwire_file_path and self.MMSTA_file_path: 
            self.process_button.config(state=tk.NORMAL)
    
    def select_output_dir(self):
        output_dir = filedialog.askdirectory()
        if output_dir:
            self.output_dir = output_dir
            self.output_label.config(text=f"Output folder: {os.path.basename(self.output_dir)}")
            if self.MMSTA_file_path and self.Maxwire_file_path and self.YPP_file_path:
                self.process_button.config(state=tk.NORMAL)
        else:   
            self.output_dir = ""
            self.output_label.config(text="Output folder: Not selected")
            self.process_button.config(state=tk.DISABLED)
    
    def check_queue(self):
        try:
            while not self.status_queue.empty():
                message = self.status_queue.get(0)
                
                if message[0] == "progress":
                    self.progress_label.config(text=f"Processing: {message[1]}")
                elif message[0] == "complete":
                    success = message[1]
                    output_file = message[2]
                    self.processing = False
                    self.process_button.config(state="normal")
                    
                    if success:
                        self.status_label.config(text=f"Success! File saved as {os.path.basename(output_file)}", fg="green")
                        self.progress_label.config(text="")
                        messagebox.showinfo("Success", "File processing completed successfully!")
                    else:
                        error = message[3] if len(message) > 3 else "Unknown error"
                        self.status_label.config(text=f"Processing failed: {error}", fg="red")
                        self.progress_label.config(text="")
                elif message[0] in ["header", "schema_no_color", "schema_color", "last_table"]:
                    success = message[1]
                    step_name = {
                        "header": "Creating header",
                        "schema_no_color": "Creating base schema",
                        "schema_color": "Adding colors",
                        "last_table": "Adding final table"
                    }
                    if success:
                        self.progress_label.config(text=f"Completed: {step_name[message[0]]}")
                    else:
                        error = message[2] if len(message) > 2 else "Unknown error"
                        self.progress_label.config(text=f"Failed: {step_name[message[0]]} - {error}")
        except Exception as e:
            print(f"Error in check_queue: {e}")
        
        # Check again after 100ms
        self.root.after(100, self.check_queue)
    
    def process_thread(self,MMSTA_file_path, Maxwire_file_path, YPP_file_path, output_file):
        try:
            if MMSTA_file_path and output_file:
                self.status_queue.put(("progress", "Loading MMSTA data..."))
                try:
                    df_mmsta = pd.read_excel(MMSTA_file_path)
                    df_mmsta = df_mmsta[['Level', 'Material Type' , 'Material', 'Material Description', 'Component quantity']]
                    
                    df_mmsta = add_columns(df_mmsta)
                    df_mmsta = Fill('0', df_mmsta)
                    df_mmsta = delete_YFG_YDM(df_mmsta)
                    df_mmsta.reset_index(drop=True, inplace=True)

                    for level in df_mmsta['Level'].unique().tolist():
                        sn = 'SN'+ level[-1]
                        des = 'DES' + level[-1]
                        df_mmsta = Fill_with_col(level, [sn, des], df_mmsta)
                    
                    df_mmsta = df_mmsta[['Level' ,'PN','SN1','DES1','SN2','DES2','SN3','DES3', 'Component quantity']]

                    df_mmsta_pivot = pd.pivot_table(df_mmsta, values = 'Component quantity', index = ['Level','SN1','DES1','SN2' , 'DES2','SN3', 'DES3'], 
                                        columns = 'PN', aggfunc = 'size', fill_value = ' ')
                    df_mmsta_pivot = df_mmsta_pivot.reset_index()
                    
                    app = xw.App(visible=False)  # Create an invisible Excel application
                    wb = app.books.add()
                    sheet_mmsta = wb.sheets[0]
                    sheet_mmsta.name = "sep mmsta"

                    
                    df_mmsta_pivot.loc[df_mmsta_pivot['Level'] == '*1', ['SN2', 'DES2', 'SN3', 'DES3']] = '(blank)'
                    df_mmsta_pivot.loc[df_mmsta_pivot['Level'] == '**2', ['SN3', 'DES3']] = '(blank)'        
                    
                    df_mmsta_pivot[df_mmsta_pivot.columns[7:]] = df_mmsta_pivot[df_mmsta_pivot.columns[7:]].apply(pd.to_numeric, errors='coerce').fillna(0).astype(int)
                    
                    columns_to_agg = list(df_mmsta_pivot.columns[7:])
                    cols = ['Level', 'SN1', 'DES1', 'SN2', 'DES2', 'SN3', 'DES3']
                    cols.extend(columns_to_agg)
                    
                    agg_dict = {col: 'sum' for col in columns_to_agg}

                    df_mmsta_pivot = df_mmsta_pivot.groupby(['Level','SN1','SN2','SN3']).agg({
                        'DES1': 'first',
                        'DES2': 'first',
                        'DES3': 'first',
                        **agg_dict
                        
                    }).reset_index()

                    df_mmsta_pivot = df_mmsta_pivot[cols]
                    df_mmsta_pivot.replace(to_replace=" " , value="(Blank)")
                    df_mmsta_pivot.replace(to_replace= 0 , value=" ")


                    sheet_mmsta.range("A1").value = [df_mmsta_pivot.columns.tolist()] + df_mmsta_pivot.values.tolist()
                    
                    ## Fil Simple
                    df_fil_simple = df_mmsta_pivot [df_mmsta_pivot['DES1'].apply(lambda x : 'circuit' in x.lower())]
                    columns_to_agg = list(df_fil_simple.columns[7:])

                    agg_dict = {col: 'sum' for col in columns_to_agg}

                    df_fil_simple = df_fil_simple.groupby(['Level','SN1']).agg({
                        'DES1': 'first',
                        'SN2': 'first',
                        'DES2': 'first',
                        'SN3': 'first',
                        'DES3': 'first',
                        **agg_dict
                        
                    }).reset_index()

                    df_fil_simple.replace(to_replace=" " , value="(Blank)")

                    fil_simple_sheet = wb.sheets.add(after=sheet_mmsta)
                    fil_simple_sheet.name = "Fil Simple"
                    fil_simple_sheet.range("A1").value = [df_fil_simple.columns.tolist()] + df_fil_simple.values.tolist()
                    
                    # Fil Double
                    df_double = df_mmsta_pivot [df_mmsta_pivot['DES1'].apply(lambda x : 'double' in x.lower())]
                    df_double = df_double.groupby(['Level','SN1']).agg({
                        'DES1': 'first',
                        'SN2': 'first',
                        'DES2': 'first',
                        'SN3': 'first',
                        'DES3': 'first',
                        **agg_dict
                        
                    }).reset_index()

                    double_sheet = wb.sheets.add(after=fil_simple_sheet)
                    double_sheet.name = "Double"
                    double_sheet.range("A1").value = [df_double.columns.tolist()] + df_double.values.tolist()
                    
                    # Fil Twisted
                    df_twisted = df_mmsta_pivot [df_mmsta_pivot['DES1'].apply(lambda x : 'twisted' in x.lower())]
                    df_twisted = df_twisted.groupby(['Level','SN1','SN2','SN3']).agg({
                        'DES1': 'first',
                        'DES2': 'first',
                        'DES3': 'first',
                        **agg_dict
                        
                    }).reset_index()
                    twisted_sheet = wb.sheets.add(after=double_sheet)
                    twisted_sheet.name = "Twisted"
                    twisted_sheet.range("A1").value = [df_twisted.columns.tolist()] + df_twisted.values.tolist()
                    
                    # Fil Joint
                    df_joint = df_mmsta_pivot [df_mmsta_pivot['DES1'].apply(lambda x : 'joint' in x.lower())]
                    df_joint = df_joint.groupby(['Level','SN1','SN2','SN3']).agg({
                        'DES1': 'first',
                        'DES2': 'first',
                        'DES3': 'first',
                        **agg_dict
                        
                    }).reset_index()
                    joint_sheet = wb.sheets.add(after=twisted_sheet)
                    joint_sheet.name = "joint"
                    joint_sheet.range("A1").value = [df_joint.columns.tolist()] + df_joint.values.tolist()
                    
                    # Sup GRP
                    df_super_grp = df_mmsta_pivot [df_mmsta_pivot['DES1'].apply(lambda x : 'super group' in x.lower())]
                    df_super_grp = df_super_grp.groupby(['Level','SN1','SN2','SN3']).agg({
                        'DES1': 'first',
                        'DES2': 'first',
                        'DES3': 'first',
                        **agg_dict
                        
                    }).reset_index()
                    superGrp_sheet = wb.sheets.add(after=joint_sheet)
                    superGrp_sheet.name = "Super group"
                    superGrp_sheet.range("A1").value = [df_super_grp.columns.tolist()] + df_super_grp.values.tolist()
                    
                    #Add summary Sheet
                    df_2 = df_super_grp[df_super_grp['SN2'] != '(blank)']
                    shape_super_grp = df_2[df_2['SN3'] != '(blank)'].shape[0]

                    shape_joint = df_joint[df_joint['SN2'] != '(blank)'].shape[0]

                    shape_twisted = df_twisted[df_twisted['SN2'] != '(blank)'].shape[0]

                    df_double_filtred = df_mmsta_pivot[df_mmsta_pivot['DES1'].apply(lambda x : 'double' in x.lower())]
                    shape_double = df_double_filtred[df_double_filtred['SN2'] != '(blank)'].shape[0]

                    shape_simple = df_fil_simple.shape[0]


                    summary = wb.sheets.add(after=superGrp_sheet)
                    summary.name = "Summary"
                    summary.range("A1").value = 'TYPE'
                    summary.range("B1").value = 'Count df_mmsta'
                    summary.range("A2").value = 'Circuit'
                    summary.range("B2").value = df_mmsta_pivot[df_mmsta_pivot['DES1'].apply(lambda x : 'circuit' in x.lower())].shape[0]
                    summary.range("A3").value = 'Double'
                    summary.range("B3").value = shape_double
                    summary.range("A4").value = 'Twisted'
                    summary.range("B4").value = shape_twisted
                    summary.range("A5").value = 'Joint'
                    summary.range("B5").value = shape_joint
                    summary.range("A6").value = 'Super group'
                    summary.range("B6").value = shape_super_grp
                    
                    # timestamp = time.strftime("%H_%M_%S")
                    # base_output_path = os.path.join(output_file, f"MMSTA_Processed_{timestamp}")
                    # output_mmsta_file = f"{base_output_path}.xlsx"
                    timestamp = time.strftime("%Y%m%d-%H_%M_%S")
                    output_file_mmsta = os.path.join(output_file, f"MMSTA_Processed_{timestamp}.xlsx")
                    print(output_file_mmsta)
                    wb.save(output_file_mmsta)
                    wb.close()
                    
                    self.status_queue.put(("progress", f"MMSTA file processed and saved as {os.path.basename(output_file_mmsta)}"))

                except Exception as e:
                    self.status_queue.put(("complete", False, output_file_mmsta, f"Failed to process MMSTA file: {str(e)}"))
                    return
           
           
            if Maxwire_file_path:
                self.status_queue.put(("progress", "Loading Maxwire List data..."))
                try:
                    df_wire_list = pd.read_excel(Maxwire_file_path)
                    
                    loc = df_wire_list.columns.get_loc('Wire Internal Name')
                    df_wire_list.insert(loc + 1, 'DES', None, allow_duplicates=False)
                    df_wire_list.insert(loc + 1, 'Sup Grp', None, allow_duplicates=False)
                    df_wire_list.insert(loc + 1, 'SN2', None, allow_duplicates=False)
                    df_wire_list.insert(loc + 1, 'SN1', None, allow_duplicates=False)
                    
                    PN_columns_wirelist = df_wire_list.columns[58:]
                    PN_columns_wirelist = list(PN_columns_wirelist)
                    PN_columns_wirelist = [col.split(':')[0] for col in PN_columns_wirelist]

                    num_cols_to_rename = len(df_wire_list.columns[58:])
                    df_wire_list.columns = df_wire_list.columns[:58].tolist() + PN_columns_wirelist


                    df_wire_list['Wire Internal Name'] = df_wire_list['Wire Internal Name'].astype(str).str.replace('W', '').astype(int)

                    df_wire_list[PN_columns_wirelist] = df_wire_list[PN_columns_wirelist].replace('X', 1)
                    df_wire_list[PN_columns_wirelist] = df_wire_list[PN_columns_wirelist].apply(pd.to_numeric, errors='coerce').fillna(0).astype(int)
                    
                    if YPP_file_path:
                        self.status_queue.put(("progress", "Loading YPP data..."))
                        try:
                            df_ypp_cae = pd.read_excel(YPP_file_path)
                            df_ypp_cae = df_ypp_cae[df_ypp_cae['Product number'].apply(lambda x : x in PN_columns_wirelist)]

                            df_ypp_cae_pivot = pd.pivot_table(df_ypp_cae, values = 'Counter',index = ['Material','Composition No'],
                                columns = 'Product number', aggfunc = 'count', fill_value = ' ')
                            
                            df_ypp_cae_pivot.reset_index(inplace=True)
                            df_ypp_cae_pivot[df_ypp_cae_pivot.columns[2:]] = df_ypp_cae_pivot[df_ypp_cae_pivot.columns[2:]].apply(pd.to_numeric, errors='coerce').fillna(0).astype(int)
                            
                            PN_columns = df_ypp_cae_pivot.columns[2:]
                            PN_columns = list(PN_columns)
                            
                            duplicate_materials = df_ypp_cae_pivot['Material'].value_counts()
                            duplicates = duplicate_materials[duplicate_materials > 1]
                            
                            
                            mmsta_sep = pd.read_excel(output_file_mmsta, sheet_name='sep mmsta')
                            mmsta_sep = Add_TYPE_col(mmsta_sep)
                            mmsta_sep[PN_columns] = mmsta_sep[PN_columns].apply(pd.to_numeric, errors='coerce').fillna(0).astype(int)
                            
                            
                            material_dict = []
                            sn_duplicated = check_duplicate(df_ypp_cae_pivot)

                            for sn in sn_duplicated:
                                valeur_a_chercher = sn
                                colonnes_trouvees = [col for col in mmsta_sep.columns if valeur_a_chercher in mmsta_sep[col].values][0]
                                mmsta_filtred_by_sn =  mmsta_sep[mmsta_sep[colonnes_trouvees] == valeur_a_chercher]
                                ypp_filtred_by_sn = df_ypp_cae_pivot[df_ypp_cae_pivot['Material'] == sn]
                                indices_valides = get_idx_duplicate(df_ypp_cae_pivot[df_ypp_cae_pivot['Material']== sn] ,mmsta_sep[mmsta_sep[colonnes_trouvees] == sn], PN_columns)
                                material_dict.extend(indices_valides)
                                #drop ypp_filtred_by_sn dans ypp
                                df_ypp_cae_pivot = df_ypp_cae_pivot[~df_ypp_cae_pivot['Material'].isin(sn_duplicated)]

                            sn_duplicated = check_duplicate(df_ypp_cae_pivot)

                            if len(sn_duplicated) == 0:
                                material_to_comp_dict = dict(zip(df_ypp_cae_pivot['Material'], df_ypp_cae_pivot['Composition No']))
                                list(material_to_comp_dict.items())[:5]   
                            
                            
                            mmsta_sep['Wire Internal Name'] = None
                            for matching in material_dict:
                                mmsta_sep.loc[list(matching[1]), 'Wire Internal Name'] = matching[0]
                                
                                
                            idx_1 = mmsta_sep[mmsta_sep['Wire Internal Name'].isnull()].index
                            mmsta_sep.loc[idx_1, 'Wire Internal Name'] = mmsta_sep.loc[idx_1, 'SN1'].map(material_to_comp_dict)

                            idx_2 = mmsta_sep[mmsta_sep['Wire Internal Name'].isnull()].index
                            mmsta_sep.loc[idx_2, 'Wire Internal Name'] = mmsta_sep.loc[idx_2, 'SN2'].map(material_to_comp_dict)

                            idx_3 = mmsta_sep[mmsta_sep['Wire Internal Name'].isnull()].index
                            mmsta_sep.loc[idx_3, 'Wire Internal Name'] = mmsta_sep.loc[idx_3, 'SN3'].map(material_to_comp_dict)
                            
                            df_wire_list, list_idx_integrated = processing_function(df_wire_list, mmsta_sep, PN_columns)
                            df_wire_list = df_wire_list.reset_index(drop=True)   
                            
                            timestamp = time.strftime("%Y%m%d-%H_%M_%S")
                            output_file_data_integrated = os.path.join(output_file, f"Data_Integrated_{timestamp}.xlsx")
                            df_wire_list.to_excel(output_file_data_integrated, sheet_name='Integrated', index=False) 
                            
                            mmsta_not_integrated = mmsta_sep.loc[~mmsta_sep.index.isin(list_idx_integrated)].reset_index(drop=True)

                            sup_grp_values = set(df_wire_list[df_wire_list['SN1'] != '(blank)']['Sup Grp'].unique())
                            sn2_values = set(df_wire_list[df_wire_list['SN2'] != '(blank)']['SN2'].unique())
                            sn1_values = set(df_wire_list[df_wire_list['SN1'] != '(blank)']['SN1'].unique())

                            all_sn_values = sup_grp_values.union(sn2_values).union(sn1_values)
                            all_sn_values = set(all_sn_values)

                            df_updated_notIntegrated = mmsta_not_integrated[
                                ~mmsta_not_integrated['SN1'].isin(all_sn_values) &
                                ~mmsta_not_integrated['SN2'].isin(all_sn_values) &
                                ~mmsta_not_integrated['SN3'].isin(all_sn_values)
                                ]

                            with pd.ExcelWriter(output_file_data_integrated, mode='a', engine='openpyxl') as writer:
                                df_updated_notIntegrated.to_excel(writer, sheet_name='Not Integrated', index=False)  
                            
            
                            mmsta_summary = pd.read_excel(output_file_mmsta, sheet_name='Summary')
                            df_updated_Integrated = pd.read_excel(output_file_data_integrated, sheet_name='Integrated')
                            df_updated_Integrated['TYPE'].value_counts()
                            type_counts_df_wireList = df_updated_Integrated['TYPE'].value_counts().reset_index()
                            type_counts_df_wireList.columns = ['TYPE', 'Count df_MaxWireList']
                            
                            concatenated_df = pd.merge(mmsta_summary[['TYPE', 'Count df_mmsta']], 
                                                                    type_counts_df_wireList[['TYPE', 'Count df_MaxWireList']], 
                                                                    on='TYPE', 
                                                                    how='inner')
                            with pd.ExcelWriter(output_file_data_integrated, mode='a', engine='openpyxl') as writer:
                                concatenated_df.to_excel(writer, sheet_name='Summary', index=False)  
                            
                            wb = load_workbook(output_file_data_integrated)

                            sheet = wb['Summary']

                            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                            for row in range(2, sheet.max_row + 1):  
                                col1 = sheet[f'B{row}']  
                                col2 = sheet[f'C{row}'] 

                                if col1.value == col2.value:
                                    col1.fill = green_fill
                                    col2.fill = green_fill
                                elif col1.value > col2.value:
                                    col1.fill = red_fill
                                else:
                                    col2.fill = red_fill

                            wb.save(output_file_data_integrated)
                            os.startfile(output_file_data_integrated)
                            
                        except Exception as e:
                            self.status_queue.put(("complete", False, output_file, f"Failed to process YPP file: {str(e)}"))
                            return
                    
                    
                except Exception as e:
                    self.status_queue.put(("complete", False, output_file, f"Failed to process Maxwire List file: {str(e)}"))
                    return
            

           
            self.status_queue.put(("complete", True, output_file_data_integrated))
            
            
        except Exception as e:
            print(f"Error in process_thread: {e}")
            self.status_queue.put(("complete", False, output_file, str(e)))
    
    def main_separator(self):
        if not self.MMSTA_file_path and not self.output_dir :
            messagebox.showerror("Error", "Please select MMSTA file and output folder")
            return
        
        if self.processing:
            messagebox.showinfo("Info", "Processing already in progress. Please wait.")
            return
        
        try:
            
            self.processing = True
            self.process_button.config(state=tk.DISABLED)
            self.status_label.config(text="Processing...", fg="blue")
            
            processing_thread = threading.Thread(
                target=self.process_thread,
                args=( 
                    self.MMSTA_file_path, 
                    self.Maxwire_file_path, 
                    self.YPP_file_path, 
                    self.output_dir
                )
            )
            processing_thread.daemon = True
            processing_thread.start()
            
        except Exception as e:
            self.processing = False
            self.process_button.config(state="normal")
            messagebox.showerror("Error", f"Error starting processing: {e}")





if __name__ == "__main__":
    root = tk.Tk()
    root.title("Data Integrator Processor - YMM-1")
    
    # Adjust window size to accommodate the new buttons
    window_width = 500
    window_height = 550
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x_position = (screen_width - window_width) // 2
    y_position = (screen_height - window_height) // 2
    root.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
    
    app = ExcelProcessor(root)

    root.mainloop()